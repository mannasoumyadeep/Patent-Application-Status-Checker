import streamlit as st
import pandas as pd
import concurrent.futures
import os
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoAlertPresentException, TimeoutException
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import json
import time
from datetime import datetime
import threading
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
import io
import base64

# Constants
APPLICATION_STATUS_URL = "https://iprsearch.ipindia.gov.in/PublicSearch/PublicationSearch/ApplicationStatus"
CAPTCHA_URL = "https://iprsearch.ipindia.gov.in/PublicSearch/Captcha/CaptchaAudio"
MAX_WORKERS = 5
MAX_RETRIES = 3
RETRY_DELAY = 5  # seconds

# Page configuration
st.set_page_config(
    page_title="Patent Application Status Checker",
    page_icon="🔍",
    layout="wide"
)

# Initialize session state
if 'processing' not in st.session_state:
    st.session_state.processing = False
if 'results' not in st.session_state:
    st.session_state.results = []
if 'errors' not in st.session_state:
    st.session_state.errors = []
if 'stop_requested' not in st.session_state:
    st.session_state.stop_requested = False
if 'progress' not in st.session_state:
    st.session_state.progress = 0
if 'total_apps' not in st.session_state:
    st.session_state.total_apps = 0
if 'application_numbers' not in st.session_state:
    st.session_state.application_numbers = []
if 'retry_individual' not in st.session_state:
    st.session_state.retry_individual = []

def parse_date(date_string):
    if not date_string:
        return None
    try:
        return datetime.strptime(date_string, "%d/%m/%Y").date().strftime("%d/%m/%Y")
    except ValueError:
        return date_string

class ApplicationService:
    def __init__(self):
        self.processed_data = {}
        self.error_applications = []
        self.stop_event = threading.Event()
        self.progress_lock = threading.Lock()
        self.total_applications = 0
        self.processed_applications = 0

    def setup_driver(self):
        options = Options()
        options.add_argument("--headless")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        options.page_load_strategy = 'eager'
        
        # Additional options for Streamlit Cloud
        options.add_argument("--disable-setuid-sandbox")
        options.add_argument("--remote-debugging-port=9222")
        options.add_argument("--disable-web-security")
        options.add_argument("--disable-features=VizDisplayCompositor")
        options.add_argument("--single-process")
        
        # Check if running on Streamlit Cloud
        import os
        if os.environ.get("STREAMLIT_SHARING_MODE"):
            # Use system Chrome on Streamlit Cloud
            options.binary_location = "/usr/bin/chromium"
            service = ChromeService(executable_path="/usr/bin/chromedriver")
        else:
            # Use ChromeDriverManager for local development
            service = ChromeService(ChromeDriverManager().install())
        
        driver = webdriver.Chrome(service=service, options=options)
        wait = WebDriverWait(driver, 15)
        return driver, wait

    def process_application_number(self, application_number, retry_count=0):
        if st.session_state.stop_requested:
            return application_number, None

        driver, wait = self.setup_driver()
        try:
            driver.get(APPLICATION_STATUS_URL)

            try:
                alert = driver.switch_to.alert
                alert.accept()
            except NoAlertPresentException:
                pass

            element1 = wait.until(EC.element_to_be_clickable((By.ID, "ApplicationNumber")))
            element1.clear()
            element1.send_keys(application_number)
            time.sleep(0.5)

            captcha_text = self.get_captcha_text(driver, wait)
            input_field1 = wait.until(EC.element_to_be_clickable((By.ID, "CaptchaText")))
            input_field1.clear()
            input_field1.send_keys(captcha_text)
            time.sleep(0.5)

            submit_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@value='Show Status']")))
            submit_button.click()

            data = self.extract_application_data(driver, wait, application_number)
            if data is None and retry_count < MAX_RETRIES:
                time.sleep(RETRY_DELAY)
                return self.process_application_number(application_number, retry_count + 1)
            return application_number, data

        except Exception as e:
            if retry_count < MAX_RETRIES:
                time.sleep(RETRY_DELAY)
                return self.process_application_number(application_number, retry_count + 1)
            return application_number, None
        finally:
            driver.quit()
            if retry_count == 0:
                with self.progress_lock:
                    self.processed_applications += 1
                    st.session_state.progress = self.processed_applications / self.total_applications

    def get_captcha_text(self, driver, wait):
        driver.execute_script(f"window.open('{CAPTCHA_URL}','_blank');")
        driver.switch_to.window(driver.window_handles[-1])
        
        element2 = wait.until(EC.presence_of_element_located((By.TAG_NAME, "pre")))
        captcha = element2.text
        json_data = json.loads(captcha)
        
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        
        return json_data["CaptchaImageText"]

    def extract_application_data(self, driver, wait, application_number):
        try:
            body = wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            tables = body.find_elements(By.TAG_NAME, "table")

            if tables:
                data = {"Application Number": application_number}
                for table in tables[:2]:
                    self.extract_table_data(table, data)
                return data
            return None

        except TimeoutException:
            return None
        except Exception as e:
            return None

    def extract_table_data(self, table, data):
        rows = table.find_elements(By.TAG_NAME, "tr")
        for row in rows:
            columns = row.find_elements(By.TAG_NAME, "td")
            if len(columns) == 2:
                key = columns[0].text.strip()
                value = columns[1].text.strip()
                field_mapping = {
                    "APPLICANT NAME": "Applicant Name",
                    "APPLICATION TYPE": "Application Type",
                    "DATE OF FILING": "Date of Filing",
                    "TITLE OF INVENTION": "Title of Invention",
                    "FIELD OF INVENTION": "Field of Invention",
                    "E-MAIL (As Per Record)": "Email (As Per Record)",
                    "ADDITIONAL-EMAIL (As Per Record)": "Additional Email (As Per Record)",
                    "E-MAIL (UPDATED Online)": "Email (Updated Online)",
                    "PCT INTERNATIONAL APPLICATION NUMBER": "PCT International Application Number",
                    "PCT INTERNATIONAL FILING DATE": "PCT International Filing Date",
                    "PRIORITY DATE": "Priority Date",
                    "REQUEST FOR EXAMINATION DATE": "Request for Examination Date",
                    "PUBLICATION DATE (U/S 11A)": "Publication Date (U/S 11A)",
                    "APPLICATION STATUS": "Application Status"
                }
                
                if key in field_mapping:
                    if "DATE" in key:
                        data[field_mapping[key]] = parse_date(value)
                    else:
                        data[field_mapping[key]] = value

def create_excel_file(results, error_only=False):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Application Statuses" if not error_only else "Error Applications"

    headers = [
        "Application Number", "Applicant Name", "Application Type", "Date of Filing",
        "Title of Invention", "Field of Invention", "Email (As Per Record)",
        "Additional Email (As Per Record)", "Email (Updated Online)",
        "PCT International Application Number", "PCT International Filing Date",
        "Priority Date", "Request for Examination Date", "Publication Date (U/S 11A)",
        "Application Status"
    ]
    sheet.append(headers)

    date_style = NamedStyle(name='datetime', number_format='DD/MM/YYYY')
    date_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    date_columns = [4, 11, 12, 13, 14]

    for application_number, app_data in results:
        if app_data and not error_only:
            row = [
                app_data.get("Application Number", ""),
                app_data.get("Applicant Name", ""),
                app_data.get("Application Type", ""),
                parse_date_for_excel(app_data.get("Date of Filing", "")),
                app_data.get("Title of Invention", ""),
                app_data.get("Field of Invention", ""),
                app_data.get("Email (As Per Record)", ""),
                app_data.get("Additional Email (As Per Record)", ""),
                app_data.get("Email (Updated Online)", ""),
                app_data.get("PCT International Application Number", ""),
                parse_date_for_excel(app_data.get("PCT International Filing Date", "")),
                parse_date_for_excel(app_data.get("Priority Date", "")),
                parse_date_for_excel(app_data.get("Request for Examination Date", "")),
                parse_date_for_excel(app_data.get("Publication Date (U/S 11A)", "")),
                app_data.get("Application Status", "")
            ]
            sheet.append(row)
            
            for col in date_columns:
                if row[col-1]:
                    sheet.cell(sheet.max_row, col).style = date_style
        elif not app_data:
            row = [application_number] + ["ERROR"] * (len(headers) - 1)
            sheet.append(row)
            for cell in sheet[sheet.max_row]:
                cell.fill = red_fill

    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    return excel_file

def parse_date_for_excel(date_string):
    if not date_string:
        return None
    try:
        return datetime.strptime(date_string, "%d/%m/%Y").date()
    except ValueError:
        return date_string

def main():
    st.title("🔍 Patent Application Status Checker")
    st.markdown("Upload an Excel file containing application numbers to check their status from the Indian Patent Office.")

    # File upload
    uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx', 'xls'])

    if uploaded_file is not None:
        df = pd.read_excel(uploaded_file)
        application_numbers = df.iloc[:, 0].astype(str).tolist()
        st.session_state.application_numbers = application_numbers
        st.session_state.total_apps = len(application_numbers)
        
        st.success(f"✓ File uploaded successfully! Found {len(application_numbers)} application numbers.")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("🚀 Start Processing", type="primary", disabled=st.session_state.processing):
                st.session_state.processing = True
                st.session_state.stop_requested = False
                st.session_state.results = []
                st.session_state.errors = []
                st.session_state.progress = 0
                
        with col2:
            if st.button("⏹️ Stop Processing", disabled=not st.session_state.processing):
                st.session_state.stop_requested = True
                st.session_state.processing = False

    # Progress display
    if st.session_state.processing or st.session_state.results:
        st.markdown("---")
        progress_bar = st.progress(st.session_state.progress)
        status_text = st.empty()
        
        if st.session_state.processing:
            status_text.text(f"Processing: {int(st.session_state.progress * st.session_state.total_apps)}/{st.session_state.total_apps} applications")

    # Process applications
    if st.session_state.processing:
        service = ApplicationService()
        service.total_applications = len(application_numbers)
        
        placeholder = st.empty()
        results = []
        errors_list = []
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            future_to_application = {
                executor.submit(service.process_application_number, app_num): app_num 
                for app_num in application_numbers
            }
            
            for future in concurrent.futures.as_completed(future_to_application):
                if st.session_state.stop_requested:
                    executor.shutdown(wait=False)
                    break
                    
                app_num = future_to_application[future]
                try:
                    result = future.result()
                    results.append(result)
                    
                    if result[1] is None:
                        errors_list.append(result[0])
                        with placeholder.container():
                            st.error(f"❌ Error processing: {result[0]}")
                    else:
                        with placeholder.container():
                            st.success(f"✓ Successfully processed: {result[0]}")
                    
                    progress_bar.progress(st.session_state.progress)
                    status_text.text(f"Processing: {int(st.session_state.progress * st.session_state.total_apps)}/{st.session_state.total_apps} applications")
                    
                except Exception as exc:
                    results.append((app_num, None))
                    errors_list.append(app_num)
                    with placeholder.container():
                        st.error(f"❌ Exception for {app_num}: {str(exc)}")

        # Sort results to maintain input order
        results.sort(key=lambda x: application_numbers.index(x[0]))
        st.session_state.results = results
        st.session_state.errors = errors_list
        st.session_state.processing = False
        st.session_state.retry_individual = errors_list.copy()

    # Display results and download options
    if st.session_state.results:
        st.markdown("---")
        st.subheader("📊 Processing Summary")
        
        successful = len([r for r in st.session_state.results if r[1] is not None])
        failed = len([r for r in st.session_state.results if r[1] is None])
        
        col1, col2, col3 = st.columns(3)
        col1.metric("Total Processed", len(st.session_state.results))
        col2.metric("Successful", successful, delta=f"{successful/len(st.session_state.results)*100:.1f}%")
        col3.metric("Failed", failed, delta=f"-{failed/len(st.session_state.results)*100:.1f}%", delta_color="inverse")

        # Download buttons
        st.subheader("📥 Download Results")
        col1, col2 = st.columns(2)
        
        with col1:
            # All results
            all_results_file = create_excel_file(st.session_state.results)
            st.download_button(
                label="📄 Download All Results",
                data=all_results_file,
                file_name=f"patent_status_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        with col2:
            # Error results only
            if st.session_state.errors:
                error_results = [(app_num, None) for app_num in st.session_state.errors]
                error_file = create_excel_file(error_results, error_only=True)
                st.download_button(
                    label="❌ Download Error Report",
                    data=error_file,
                    file_name=f"patent_errors_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        # Retry failed applications
        if st.session_state.retry_individual:
            st.markdown("---")
            st.subheader("🔄 Retry Failed Applications")
            st.write(f"Found {len(st.session_state.retry_individual)} failed applications.")
            
            if st.button("🔄 Retry All Failed", type="secondary"):
                service = ApplicationService()
                service.total_applications = len(st.session_state.retry_individual)
                
                retry_placeholder = st.empty()
                retry_progress = st.progress(0)
                
                for idx, app_num in enumerate(st.session_state.retry_individual):
                    if st.session_state.stop_requested:
                        break
                        
                    with retry_placeholder.container():
                        st.info(f"Retrying: {app_num}")
                    
                    _, data = service.process_application_number(app_num)
                    
                    if data is not None:
                        # Update the results
                        for i, (num, _) in enumerate(st.session_state.results):
                            if num == app_num:
                                st.session_state.results[i] = (app_num, data)
                                break
                        
                        st.session_state.errors.remove(app_num)
                        with retry_placeholder.container():
                            st.success(f"✓ Successfully retried: {app_num}")
                    else:
                        with retry_placeholder.container():
                            st.error(f"❌ Failed again: {app_num}")
                    
                    retry_progress.progress((idx + 1) / len(st.session_state.retry_individual))
                
                st.session_state.retry_individual = [app for app in st.session_state.errors]
                st.rerun()

    # Instructions
    with st.expander("📖 Instructions"):
        st.markdown("""
        ### How to use this application:
        
        1. **Prepare your Excel file**: Create an Excel file with application numbers in the first column
        2. **Upload the file**: Click on the file uploader and select your Excel file
        3. **Start processing**: Click the "Start Processing" button to begin checking statuses
        4. **Monitor progress**: Watch the progress bar and status updates
        5. **Download results**: Once complete, download the results and error report
        6. **Retry failures**: If any applications failed, you can retry them individually
        
        ### Notes:
        - The application uses 5 concurrent workers for faster processing
        - Failed applications can be retried and will be merged with successful results
        - You can stop the process at any time and download partial results
        - Results maintain the original order from your input file
        """)

    # Footer
    st.markdown("---")
    st.markdown("Made with ❤️ using Streamlit | Patent Status Checker v1.0")

if __name__ == "__main__":
    main()